import os
from openpyxl import load_workbook
from src.colors import *

FILES_DIR = "files"
EXCEL_PATH = "files/External Partners Channels.xlsx"
SHEET_NAME = "Channels"

def promptExcelPath(default=EXCEL_PATH):
  """Prompt for the Excel filename to update, defaulting to the standard tracker.
  A bare filename is resolved inside files/; a value containing a path separator
  is used as-is. Empty input keeps the default."""
  default_name = os.path.basename(default)
  print(f"\n  {bold}Excel filename{reset} {stone}[{default_name}]{reset}: ", end="")
  entered = input().strip()
  if not entered:
    return default
  if os.path.dirname(entered):
    return entered
  return os.path.join(FILES_DIR, entered)

def getAvailableColumns(filepath=EXCEL_PATH):
  """Return a list of (column_letter, header_name) for columns D onwards."""
  wb = load_workbook(filepath)
  ws = wb[SHEET_NAME]
  columns = []
  for cell in ws[1]:
    if cell.column >= 4 and cell.value:  # D onwards
      columns.append((cell.column_letter, cell.value))
  wb.close()
  return columns

def updateExcel(results, column_letter, filepath=EXCEL_PATH):
  """
  Update the Excel file for a given column based on send results.
  Matches chat IDs from column A against the results dict {chat_id: True/False}.
  Writes 'Yes' for success, 'No' for failure.
  """
  wb = load_workbook(filepath)
  ws = wb[SHEET_NAME]
  updated = 0

  for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    cell_a = row[0].value
    if not cell_a:
      continue
    # Column A format: "-1001639374617, Polkadot (DOT) / Binance Tech, EXCHANGE"
    # Extract the chat ID (everything before the first comma)
    chat_id = str(cell_a).split(",")[0].strip()
    if chat_id in results:
      target_cell = ws[f"{column_letter}{row[0].row}"]
      target_cell.value = "Yes" if results[chat_id] else "No"
      updated += 1

  wb.save(filepath)
  wb.close()
  return updated
